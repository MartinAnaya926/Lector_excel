from openpyxl import load_workbook
from openpyxl import Workbook
from copy import copy 
import openpyxl
# Especifica la ruta de tu archivo Excel
archivo_excel = 'D:/Desktop/Martín Anaya/02 VSC/01 Input/Prueba.xlsx'

# Carga el archivo Excel
libro_excel = load_workbook(archivo_excel)

# Obtiene el nombre de la primera hoja
Nombres = libro_excel.sheetnames[0]

# Accede a la hoja por su nombre
hoja = libro_excel[Nombres]

#----------------------------------------------------------------------------------------------------------------------------
# Itera sobre las filas de la hoja
#for fila in hoja.iter_rows():
#    for celda in fila:
#        # Imprime el valor de cada celda
#        print(celda.value)
#-----------------------------------------------------------------------------------------------------------------------------
hoja = libro_excel.active

#Inicia la variable para contar las filas
filas_llenas = 0

#Iteración sobre las filas y conteo
for fila in hoja.iter_rows():
    if not all(cell.value is None for cell in fila):
        filas_llenas += 1

print("Número de filas llenas:", filas_llenas)
print()

#Llenar las celdas manualmente y guardar en un libro y ubicación nuevos
hoja["A2"].value = "1"
hoja["B2"].value = "2"
hoja["C2"].value = "3"

#Define ruta del archivo de salida
ruta_nuevo_archivo = "D:/Desktop/Martín Anaya/02 VSC/02 Output/Prueba_Salida.xlsx"

#Crea el nuevo archivo
Prueba_Salida = Workbook()
nueva_hoja = Prueba_Salida.active

#Copia los datos al nuevo archivo
for fila, row in enumerate(hoja.iter_rows(), start=1):
    for columna, celda in enumerate(row, start=1):
        nueva_hoja.cell(row=fila, column=columna).value = celda.value

#Guarda el nuevo archivo
Prueba_Salida.save(ruta_nuevo_archivo)

#Cierra ambos libros de excel
libro_excel.close()
Prueba_Salida.close()


