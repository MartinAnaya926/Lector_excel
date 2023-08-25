from openpyxl import load_workbook
import xlwings as xw
import matplotlib.pyplot as plt
import locale

#################################
#################################
#################################
# Set to German locale to get comma decimal separater
locale.setlocale(locale.LC_NUMERIC, 'es_ES')

plt.rcdefaults()

# Tell matplotlib to use the locale we set above
plt.rcParams['axes.formatter.use_locale'] = True
########################
while True:
    
    '''
    RUTAS QUE REQUIEREN SER MODIFICADAS
        Unicamente edite la información de las siguientes lineas de acuerdo
        a las instrucciones de uso del código
    '''

    # Ruta del archivo Excel original
    ruta_archivo_original = r"C:\Users\SHI-PC34.SHI-PC34\Desktop\Martin_Anaya\01_Aforos_Liquidos\04_Procesamiento\Formato_Aforos - propuesta.xlsx"
    # Ruta y nombre de la carpeta donde se almacenan los archivos de Excel modificados
    ruta_archivo_modificado = r"C:\Users\SHI-PC34.SHI-PC34\Desktop\Martin_Anaya\01_Aforos_Liquidos\02_VSC\02_Output\Nuevo formato"
    # Ruta del archivo que contiene los datos para el diccionario
    ruta_archivo_fuente = r"C:\Users\SHI-PC34.SHI-PC34\Desktop\Martin_Anaya\01_Aforos_Liquidos\03_Informacion_Base\Solicitud_033.xlsx"
    #Ruta para guardar las gráficas
    ruta_grafica = r"C:\Users\SHI-PC34.SHI-PC34\Desktop\Martin_Anaya\01_Aforos_Liquidos\02_VSC\02_Output\Nuevo formato"
    #Ruta imagen para el relleno de la gráfica
    textura_1 = plt.imread(r"C:\Users\SHI-PC34.SHI-PC34\Desktop\Martin_Anaya\01_Aforos_Liquidos\02_VSC\01_Input\textura-puntos.jpg")

    #################################
    #################################
    #################################

    '''
    ARCHIVO DE PROCESAMIENTO
    NO MODIFICAR
        Por favor no modifique el código presentado a continuación, en caso de
        presentarse algún error, informar acerca de este para realizar la correspondiente verificación
    '''
    
    if ruta_archivo_original.split('.')[-1] != 'xlsx':
        print('VERIFIQUE LA RUTA DE: ruta_archivo_original')
        break
    
    if ruta_archivo_fuente.split('.')[-1] != 'xlsx':
        print('VERIFIQUE LA RUTA DE: ruta_archivo_fuente')
        break
        
    if ruta_archivo_fuente.split('.')[-1] != 'xlsx':
        print('VERIFIQUE LA RUTA DE: ruta_archivo_fuente')
        break
    
    #Completar rutas
    ruta_archivo_modificado = ruta_archivo_modificado + r"\Aforo_"
    ruta_grafica = ruta_grafica + r"\Perfil de Profundidad y Velocidad_"

    # Cargar el archivo Excel original
    libro_original = load_workbook(ruta_archivo_original)
    # Leer archivo fuente
    libro_fuente = load_workbook(ruta_archivo_fuente)

    # Modificar archivo fuente
    hoja_fuente = libro_fuente.active 

    #Contador para el número de filas llenas
    filas_llenas = 0

    #Iteración sobre las y conteo de las filas llenas
    for fila in hoja_fuente.iter_rows():
        if not all(cell is None for cell in fila):
            filas_llenas += 1
    print()
    print("Número de estaciones de aforo:", filas_llenas - 1)
    print("Desde la número 2 hasta la número", filas_llenas)
    print()

    #Llenar las celdas manualmente y guardar en un libro y ubicación nuevos
    #datos son los valores que leo de un excel
    #Celdas son las celdas en las que almaceno los datos leídos en el formato que deseo modificar
    datos = {"NomProy":"C","CentCostos":"D","Corriente":"W","Fecha":"H","Estación/cod":"K","NomPunto":"K","HoraInicio":"I","HoraFin":"DJ","MiraInicio":"U","MiraFin":"V","Dpto":"M","Mpio":"N","Responsables":"X","AnchoT":"L","Tipo aforo":"Z","Correntómetro":"Y","Volumen 1":"AC","Tiempo 1":"AD","Volumen 2":"AE","Tiempo 2":"AF","Volumen 3":"AG","Tiempo 3":"AH","Abscisa 1":"AK","Prof 1":"AL","0,8H (1)":"AM","0,6H (1)":"AN","0,2H (1)":"AO","Abscisa 2":"AP","Prof 2":"AQ","0,8H (2)":"AR","0,6H (2)":"AS","0,2H (2)":"AT","Abscisa 3":"AU","Prof 3":"AV","0,8H (3)":"AW","0,6H (3)":"AX","0,2H (3)":"AY","Abscisa 4":"AZ","Prof 4":"BA","0,8H (4)":"BB","0,6H (4)":"BC","0,2H (4)":"BD","Abscisa 5":"BE","Prof 5":"BF","0,8H (5)":"BG","0,6H (5)":"BH","0,2H (5)":"BI","Abscisa 6":"BJ","Prof 6":"BK","0,8H (6)":"BL","0,6H (6)":"BM","0,2H (6)":"BN","Abscisa 7":"BO","Prof 7":"BP","0,8H (7)":"BQ","0,6H (7)":"BR","0,2H (7)":"BS","Abscisa 8":"BT","Prof 8":"BU","0,8H (8)":"BV","0,6H (8)":"BW","0,2H (8)":"BX","Abscisa 9":"BY","Prof 9":"BZ","0,8H (9)" :"CA","0,6H (9)":"CB","0,2H (9)":"CC","Abscisa 10":"CD","Prof 10":"CE","0,8H (10)":"CF","0,6H (10)":"CG","0,2H (10)":"CH","Abscisa 11":"CI","Prof 11":"CJ","0,8H (11)":"CK","0,6H (11)":"CL","0,2H (11)":"CM","Abscisa 12":"CN","Prof 12":"CO","0,8H (12)":"CP","0,6H (12)":"CQ","0,2H (12)":"CR","Abscisa 13":"CS","Prof 13":"CT","0,8H (13)":"CU","0,6H (13)":"CV","0,2H (13)":"CW","Abscisa 14":"CX","Prof 14":"CY","0,8H (14)":"CZ","0,6H (14)":"DA","0,2H (14)":"DB","Coordenada_x":"T","Coordenada_y":"T","Margen inicio":"AJ"}
    celdas = ['D4',"J4","D5",'D6','D7','D8','D9','D10','G5','G6','G7','G8','M4','J9','J6','J7','N14','O14','N19','O19','N23','O23','A14','B14','C14','D14','E14','A15','B15','C15','D15','E15','A16','B16','C16','D16','E16','A17','B17','C17','D17','E17','A18','B18','C18','D18','E18','A19','B19','C19','D19','E19','A20','B20','C20','D20','E20','A21','B21','C21','D21','E21','A22','B22','C22','D22','E22','A23','B23','C23','D23','E23','A24','B24','C24','D24','E24','A25','B25','C25','D25','E25','A26','B26','C26','D26','E26','A27','B27','C27','D27','E27','G10','G9','J5']
    datos_type = {"NomProy":'str',"CentCostos":'str',"Corriente":'str',"Fecha":'str',"Estación/cod":'str',"NomPunto":'str',"HoraInicio":"str","HoraFin":"str","MiraInicio":"float","MiraFin":"float","Dpto":'str',"Mpio":'str',"Responsables":'str',"AnchoT":'float',"Tipo aforo":"str","Correntómetro":"str","Volumen 1":"float","Tiempo 1":"float","Volumen 2":"float","Tiempo 2":"float","Volumen 3":"float","Tiempo 3":"float","Abscisa 1":'float',"Prof 1":'float',"0,8H (1)":'float',"0,6H (1)":'float',"0,2H (1)":'float',"Abscisa 2":'float',"Prof 2":'float',"0,8H (2)":'float',"0,6H (2)":'float',"0,2H (2)":'float',"Abscisa 3":'float',"Prof 3":'float',"0,8H (3)":'float',"0,6H (3)":'float',"0,2H (3)":'float',"Abscisa 4":'float',"Prof 4":'float',"0,8H (4)":'float',"0,6H (4)":'float',"0,2H (4)":'float',"Abscisa 5":'float',"Prof 5":'float',"0,8H (5)":'float',"0,6H (5)":'float',"0,2H (5)":'float',"Abscisa 6":'float',"Prof 6":'float',"0,8H (6)":'float',"0,6H (6)":'float',"0,2H (6)":'float',"Abscisa 7":'float',"Prof 7":'float',"0,8H (7)":'float',"0,6H (7)":'float',"0,2H (7)":'float',"Abscisa 8":'float',"Prof 8":'float',"0,8H (8)":'float',"0,6H (8)":'float',"0,2H (8)":'float',"Abscisa 9":'float',"Prof 9":'float',"0,8H (9)" :'float',"0,6H (9)":'float',"0,2H (9)":'float',"Abscisa 10":'float',"Prof 10":'float',"0,8H (10)":'float',"0,6H (10)":'float',"0,2H (10)":'float',"Abscisa 11":'float',"Prof 11":'float',"0,8H (11)":'float',"0,6H (11)":'float',"0,2H (11)":'float',"Abscisa 12":'float',"Prof 12":'float',"0,8H (12)":'float',"0,6H (12)":'float',"0,2H (12)":'float',"Abscisa 13":'float',"Prof 13":'float',"0,8H (13)":'float',"0,6H (13)":'float',"0,2H (13)":'float',"Abscisa 14":'float',"Prof 14":'float',"0,8H (14)":'float',"0,6H (14)":'float',"0,2H (14)":'float',"Coordenada_x":"str", "Coordenada_y":"str","Margen inicio":"str"}
    datos2 = {"AnchoT":"J9",'Abscisa 1':'A14','Profundidad 1':'B14','Velocidad media 1':'I14','Abscisa 2':'A15','Profundidad 2':'B15','Velocidad media 2':'I15','Abscisa 3':'A16','Profundidad 3':'B16','Velocidad media 3':'I16','Abscisa 4':'A17','Profundidad 4':'B17','Velocidad media 4':'I17','Abscisa 5':'A18','Profundidad 5':'B18','Velocidad media 5':'I18','Abscisa 6':'A19','Profundidad 6':'B19','Velocidad media 6':'I19','Abscisa 7':'A20','Profundidad 7':'B20','Velocidad media 7':'I20','Abscisa 8':'A21','Profundidad 8':'B21','Velocidad media 8':'I21','Abscisa 9':'A22','Profundidad 9':'B22','Velocidad media 9':'I22','Abscisa 10':'A23','Profundidad 10':'B23','Velocidad media 10':'I23','Abscisa 11':'A24','Profundidad 11':'B24','Velocidad media 11':'I24','Abscisa 12':'A25','Profundidad 12':'B25','Velocidad media 12':'I25','Abscisa 13':'A26','Profundidad 13':'B26','Velocidad media 13':'I26','Abscisa 14':'A27','Profundidad 14':'B27','Velocidad media 14':'I27'}
    #print(len(datos))
    #print(len(celdas))
    #print(len(datos_type))

    nombre = 'llenado_'

    for fila in range(2,filas_llenas + 1):
        
        try:
            Abs = []
            Prof = []
            Vel = []
            # Crear una copia del archivo original
            libro_modificado = load_workbook(ruta_archivo_original)

            # Modificar el archivo copiado
            nombre_hoja_modificada = 'Cálculos de aforo'
            hoja_modificada = libro_modificado.get_sheet_by_name(nombre_hoja_modificada)
            hoja_modificada = libro_modificado.active 

            #Dar los datos a las celdas
            lista_keys = list(datos.keys())             #Extraer una lista con las "claves" del diccionario
            for dato in range(len(datos)):
            
                key = lista_keys[dato]                  #Definir cada una de las claves, posición dato de la lista lista_keys
                pos_celda = datos[key] + str(fila)      #Obtener la posición de la celda y cocnatenar con el número de la fila
                valor = hoja_fuente[pos_celda].value
                
                if key == "NomPunto":
                    NomPunto = valor
                    print()
                    print("Punto número",str(fila)+":",NomPunto)
                    
                if key == "Margen inicio":
                    Margen = valor

                #Cambiar revoluciones en cero por un espacio vacío - INICIO
                if key == "0,8H ("+str(fila-1)+")":
                    try:
                        r1 = ".".join(str(valor).split(','))
                        r1 = float(r1)

                        if r1 == 0.0:
                            valor = None
                    except:
                        0
               
                #Cambiar revoluciones en cero por un espacio vacío - FIN
                #Este condicional convierte los datos que se requieran a flotantes
                if datos_type[key] == 'float':
                    try:
                        valor = float(valor)            #4.3 o 41,35 -- ['41','35']  -- 41.35
                    except:
                        try:
                            valor = valor.split(',')
                            valor = float(".".join(valor))
                        except:
                            valor = ""
                #Avisar si el aforo fue por vadeo o volumétrico
                if key == "Tipo aforo":
                    tipo_aforo = valor
                    print("Tipo de aforo en el punto",str(fila),"("+NomPunto+"):",tipo_aforo)

                #Los dos siguientes condicionales separan y añaden las coordenadas por separado
                if key == "Coordenada_x":
                    try:
                        valor = valor.split(" ")[0]
                    except:
                        print('No hay coordenada X en el punto', str(fila))

                if key == "Coordenada_y":
                    try:
                        valor = valor.split(" ")[1]
                    except:
                        print('No hay coordenada Y en el punto', str(fila))    
          
                hoja_modificada[celdas[dato]] = valor

            # Guardar el archivo Excel modificado en la nueva ruta
            ruta_ultima = ruta_archivo_modificado+nombre+str(fila)+"_"+NomPunto+".xlsx"
            libro_modificado.save(ruta_ultima)
            libro_modificado.close()
            
            #PARA LA GRÁFICA - PARA LA GRÁFICA - PARA LA GRÁFICA - PARA LA GRÁFICA - PARA LA GRÁFICA - PARA LA GRÁFICA - PARA LA GRÁFICA - PARA LA GRÁFICA - PARA LA GRÁFICA
            
            libro_modificado2 = xw.Book(ruta_ultima)
            app = xw.apps.active
            hoja_modificada2 = libro_modificado2.sheets.active
            
            lista_keys2 = list(datos2.keys())
            
            for dato in range(len(datos2)):
                
                key2 = lista_keys2[dato]                  
                pos_celda = datos2[key2]
                valor = hoja_modificada2[pos_celda].value

                #En este ciclo se llenan las listas (Abs, Prof, Vel) para la grafica (x ,y1, y2)
                for i in range (1,15):
                     
                    abs = "Abscisa " + str(i)
                    prof = "Profundidad " + str(i)
                    vel  = "Velocidad media " + str(i)
                
                    if key2 == abs:
                        if valor is None:
                            #print("No hay", abs)

                            celda_abs = datos2.get(key2)
                            abs_2_value = hoja_modificada2[celda_abs].value
                            Abs.append(abs_2_value)
                            
                        else:
                            celda_abs = datos2.get(key2)
                            abs_2_value = hoja_modificada2[celda_abs].value
                            Abs.append(abs_2_value)
                            
                    if key2 == prof:
                        if valor is None:

                            celda_prof = datos2.get(key2)
                            prof_2_value = hoja_modificada2[celda_prof].value
                            Prof.append(prof_2_value)

                        else:
                            celda_prof = datos2.get(key2) 
                            prof_2_value = hoja_modificada2[celda_prof].value
                            Prof.append(prof_2_value)

                    if key2 == vel:
                        if valor is None:

                            celda_vel = datos2.get(key2)
                            vel_2_value = hoja_modificada2[celda_vel].value 
                            Vel.append(vel_2_value)

                        else:
                            celda_vel = datos2.get(key2) 
                            vel_2_value = hoja_modificada2[celda_vel].value 
                            Vel.append(vel_2_value)

            #print(Abs)
            #print(Prof)
            #print(Vel)
            #print()
            #print(len(Abs))
            #print(len(Prof))
            #print(len(Vel))
            
            #Los siguientes 3 ciclos corrigen los datos almacenados en las listas
            for n in reversed(range(len(Abs))):
                if Abs[n] == None and n>0:
                    if Abs[n-1] != None:
                        Abs[n] = Abs[n-1]
                    else:
                        Abs[n] = 0.0

                elif Abs[n] == None:
                    Abs[n] = 0.0
                    
            for p in reversed(range(len(Prof))):
                if Prof[p] == None:
                    Prof[p] = 0.0
                    
            for s in reversed(range(len(Vel))):
                if Vel[s] == None:
                    Vel[s] = 0.0
                elif Vel[s] == ' ':
                    Vel[s] = 0.0 
                    
            #print()               
            #print(Abs)
            #print(Prof)
            #print(Vel)
            #print()
            #print(len(Abs))
            #print(len(Prof))
            #print(len(Vel))
            
            #Generar la gráfica y sus propiedades
            ancho13 = (max(Abs)/13)*0.5
            fig, ax1 = plt.subplots(figsize=(12,7))
            ax2 = ax1.twinx()
            fig.text(0.13, 0.01, Margen, size = 16, fontweight='bold')
            ax1.plot(Abs, Prof)
            ax2.bar(Abs,Vel, color = 'orange', edgecolor = 'brown', linewidth = 1.5, alpha = 0.6, label = 'Velocidad', width = ancho13)
            ax1.fill_between(Abs, Prof, color='skyblue')
            ax2.invert_yaxis() 
            ax1.invert_yaxis()
            ax1.set_xticks(Abs)
            ax1.tick_params(axis = 'x', labelsize = 16)
            ax1.tick_params(axis = 'y', labelsize = 16)
            ax2.tick_params(axis = 'y', labelsize = 16)
            
            ax1.imshow(textura_1, extent=[ax1.get_xlim()[0], ax1.get_xlim()[1], ax1.get_ylim()[0], ax1.get_ylim()[1]], aspect='auto', alpha=0.5, cmap='gray', interpolation='bilinear')

            titulo = 'Perfil de Profundidad y Velocidad '+str(NomPunto)
            ax1.set_title(titulo, fontweight='bold', fontsize=18, pad = 20)
            ax1.set_xlabel('Ancho (m)', fontsize=16, labelpad = 15)
            ax1.set_ylabel('Profundidad (m)', fontsize=16, labelpad = 15)
            ax2.set_ylabel('Velocidad del Flujo (m/s)', fontsize=16, labelpad = 15)

            plt.legend()
            
            ruta_ultima_grafica = ruta_grafica+NomPunto+".png"
            plt.savefig(ruta_ultima_grafica, dpi = 600)
            
            imagen = hoja_modificada2.pictures.add(ruta_ultima_grafica, left=hoja_modificada2.range('B33').left, top=hoja_modificada2.range('B33').top)
            #imagen.width = 600
            #imagen.height = 500
            #plt.show()
            plt.close()
            libro_modificado2.save()
            libro_modificado2.close()
            
            app.kill()
        
        except:
            print("########### ALERTA")
            print("  No es posible procesar el archivo")
            print("  Ruta del archivo: ", ruta_archivo_original)
            print("  Fila: ", fila)
        
    # Cerrar los archivos
    libro_original.close()
    break
