from openpyxl import load_workbook

# Especifica la ruta de tu archivo Excel
archivo_excel = 'r.E:\Martín Anaya\01 Aforos Líquidos\02 VSC\01 Input\Prueba.xlsx'

# Carga el archivo Excel
libro_excel = load_workbook(archivo_excel)

# Obtiene el nombre de la primera hoja
Nombres = libro_excel.sheetnames[0]

# Accede a la hoja por su nombre
hoja = libro_excel[Nombres]

# Itera sobre las filas de la hoja
for fila in hoja.iter_rows():
    for celda in fila:
        # Imprime el valor de cada celda
        print(celda.value)
