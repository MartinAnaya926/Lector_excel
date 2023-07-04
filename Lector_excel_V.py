from openpyxl import load_workbook
import openpyxl

# Especifica la ruta de tu archivo Excel
archivo_excel = 'D:/Desktop/Martín Anaya/03 Información Base/Solicitud_033.xlsx'

# Carga el archivo Excel
libro_excel = load_workbook(archivo_excel)

# Obtiene el nombre de la primera hoja
Nombres = libro_excel.sheetnames[0]

# Accede a la hoja por su nombre
hoja = libro_excel[Nombres]

# Itera sobre las filas de la hoja
#for fila in hoja.iter_rows():
#    for celda in fila:
#        # Imprime el valor de cada celda
#        print(celda.value)

hoja = libro_excel.active

#Inicia la variable para contar las filas
filas_llenas = 0

#Iteración sobre las filas y conteo
for fila in hoja.iter_rows():
    if not all(cell.value is None for cell in fila):
        filas_llenas += 1

print("Número de filas llenas:", filas_llenas)

#TAREA: LLenas variables con la columna correspondiente
datos = {'NomProy':'C','CentCostos':'D','Corriente':'W',"Fecha":'H','Estación/cod':'K','NomPunto':'K','Dpto':'M','Mpio':'N','Responsables':'X','AnchoT':'L','Abscisa 1':'AK','Prof':'AL','0,8H':'AM','0,6H':'AN','0,2H':'AO','Abscisa 2':'AP','Prof':'AQ','0,8H':'AR','0,6H':'AS','0,2H':'AT','Abscisa 3':'AU','Prof':'AV','0,8H':'AW','0,6H':'AX','0,2H':'AY','Abscisa 4':'AZ','Prof':'BA','0,8H':'BB','0,6H':'BC','0,2H':'BD','Abscisa 5':'BE','Prof':'BF','0,8H':'BG','0,6H':'BH','0,2H':'BI','Abscisa 6':'BJ','Prof':'BK','0,8H':'BL','0,6H':'BM','0,2H':'BN','Abscisa 7':'BO','Prof':'BP','0,8H':'BQ','0,6H':'BR','0,2H':'BS','Abscisa 8':'BT','Prof':'BU','0,8H':'BV','0,6H':'BW','0,2H':'BX','Abscisa 9':'BY','Prof':'BZ','0,8H':'CA','0,6H':'CB','0,2H':'CC','Abscisa 10':'CD','Prof':'CE','0,8H':'CF','0,6H':'CG','0,2H':'CH','Abscisa 11':'CI','Prof':'CJ','0,8H':'CK','0,6H':'CL','0,2H':'CM','Abscisa 12':'CN','Prof':'CO','0,8H':'CP','0,6H':'CQ','0,2H':'CR','Abscisa 13':'CS','Prof':'CT','0,8H':'CU','0,6H':'CV','0,2H':'CW','Abscisa 14':'CX','Prof':'CY','0,8H':'CZ','0,6H':'DA','0,2H':'DB'}
celdas = datos.values()

"TAREA: Que en un archivo de excel, me queden los siguientes 3 datos en las celdas A2, B2, C2"
#1. Leer un archivo de excel
#2. Editarlo, el orginal debe quedar sin ninguna modificación
#3. Guardar con un nuevo nombre

#TAREA: Ensayar
fila = 2
print()
print("DATOS LEIDOS")
print()
for k in list(datos.keys()):
    pos = datos[k]+str(fila)
    print(k,hoja[pos].value, sep=": ")

