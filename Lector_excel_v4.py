from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
# Especifica la ruta de tu archivo Excel
archivo_excel = 'D:/Desktop/Martín Anaya/02 VSC/01 Input/Prueba.xlsx'

# Carga el archivo Excel
libro_excel = load_workbook(archivo_excel)

# Obtiene el nombre de la primera hoja
Nombres = libro_excel.sheetnames[0]

# Accede a la hoja por su nombre
hoja = libro_excel[Nombres]

#----------------------------------------------------------------------------------------------------------------------------
# Itera sobre las filas de la hoja
#for fila in hoja.iter_rows():
#    for celda in fila:
#        # Imprime el valor de cada celda
#        print(celda.value)
#-----------------------------------------------------------------------------------------------------------------------------
hoja = libro_excel.active

#Inicia la variable para contar las filas
filas_llenas = 0

#Iteración sobre las filas y conteo
for fila in hoja.iter_rows():
    if not all(cell.value is None for cell in fila):
        filas_llenas += 1

print("Número de filas llenas:", filas_llenas)
print()

#Llenar celdas a partir de una lista dada "prueba"
prueba = ["Alacrán",12154656,"Río Cauca"]
celdas = ['A2',"B2","D2"]

#Dar los datos a las celdas
for dato in range(len(prueba)):
    hoja[celdas[dato]] = prueba[dato]

#Definir la ruta del archivo de salida
ruta_nuevo_archivo = "D:/Desktop/Martín Anaya/02 VSC/02 Output/Prueba_Salida_3.xlsx"

#Crear el nuevo archivo 
Prueba_Salida= Workbook()
nueva_hoja = Prueba_Salida.active

for fila_orig, fila_nueva in zip(hoja.iter_rows(), nueva_hoja.iter_rows()):
    for celda_orig, celda_nueva in zip(fila_orig, fila_nueva):
        celda_nueva.value = celda_orig.value
        celda_nueva.font = copy(celda_orig.font)
        celda_nueva.border = copy(celda_orig.border)
        celda_nueva.fill = copy(celda_orig.fill)
        celda_nueva.number_format = copy(celda_orig.number_format)
        celda_nueva.alignment = copy(celda_orig.alignment)

#Copiar los datos al nuevo archivo
for fila, row in enumerate(hoja.iter_rows(), start=1):
    for columna, celda in enumerate(row, start=1):
        nueva_hoja.cell(row=fila, column=columna).value = celda.value

#Guardar el nuevo archivo
Prueba_Salida.save(ruta_nuevo_archivo)

#Cerrar ambos libros de excel
libro_excel.close()
Prueba_Salida.close()