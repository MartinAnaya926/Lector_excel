from openpyxl import load_workbook

# Ruta del archivo Excel original
ruta_archivo_original = "D:/Desktop/Martín Anaya/04 Procesamiento/Formato_Aforos.xlsx"

# Ruta y nombre del archivo Excel modificado
ruta_archivo_modificado = "D:/Desktop/Martín Anaya/02 VSC/02 Output/Aforos_LaBermejala.xlsx"

# Cargar el archivo Excel original
libro_original = load_workbook(ruta_archivo_original)

# Crear una copia del archivo original
libro_modificado = load_workbook(ruta_archivo_original)

# Modificar el archivo copiado
hoja_modificada = libro_modificado.active 

#PRUEBA MANUAL CON EL FORMATO DE AFOROS
#Llenar las celdas manualmente y guardar en un libro y ubicación nuevos
hoja_modificada["D4"].value = "Moravia"
hoja_modificada["J4"].value = "1111111"
hoja_modificada["D5"].value = "La Bermejala"
hoja_modificada["D6"].value = "13-06-2023"
hoja_modificada["D7"].value = "Los Puentes"

# Guardar el archivo Excel modificado en la nueva ruta
libro_modificado.save(ruta_archivo_modificado)

# Cerrar los archivos
libro_original.close()
libro_modificado.close()