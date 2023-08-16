from openpyxl import load_workbook

# Ruta del archivo Excel original
ruta_archivo_original = r"C:\Users\SHI-PC34.SHI-PC34\Desktop\Martín Anaya\01 Aforos Líquidos\04 Procesamiento\Formato_Aforos.xlsx"
# Ruta y nombre del archivo Excel modificado
ruta_archivo_modificado = r"C:\Users\SHI-PC34.SHI-PC34\Desktop\Martín Anaya\01 Aforos Líquidos\02 VSC\02 Output\Aforos_"
# Ruta del archivo que contiene los datos para el diccionario
ruta_archivo_fuente = r"C:\Users\SHI-PC34.SHI-PC34\Desktop\Martín Anaya\01 Aforos Líquidos\03 Información Base\Solicitud_033.xlsx"

# Cargar el archivo Excel original
libro_original = load_workbook(ruta_archivo_original)
# Leer archivo fuente
libro_fuente = load_workbook(ruta_archivo_fuente)

# Modificar archivo fuente
hoja_fuente = libro_fuente.active 

#Llenar las celdas manualmente y guardar en un libro y ubicación nuevos
#datos son los valores que leo de un excel
#Celdas son las celdas en las que almaceno los datos leídos en el formato que deseo modificar
datos = {"NomProy":"C","CentCostos":"D","Corriente":"W","Fecha":"H","Estación/cod":"K","NomPunto":"K","Dpto":"M","Mpio":"N","Responsables":"X","AnchoT":"L","Abscisa 1":"AK","Prof 1":"AL","0,8H (1)":"AM","0,6H (1)":"AN","0,2H (1)":"AO","Abscisa 2":"AP","Prof 2":"AQ","0,8H (2)":"AR","0,6H (2)":"AS","0,2H (2)":"AT","Abscisa 3":"AU","Prof 3":"AV","0,8H (3)":"AW","0,6H (3)":"AX","0,2H (3)":"AY","Abscisa 4":"AZ","Prof 4":"BA","0,8H (4)":"BB","0,6H (4)":"BC","0,2H (4)":"BD","Abscisa 5":"BE","Prof 5":"BF","0,8H (5)":"BG","0,6H (5)":"BH","0,2H (5)":"BI","Abscisa 6":"BJ","Prof 6":"BK","0,8H (6)":"BL","0,6H (6)":"BM","0,2H (6)":"BN","Abscisa 7":"BO","Prof 7":"BP","0,8H (7)":"BQ","0,6H (7)":"BR","0,2H (7)":"BS","Abscisa 8":"BT","Prof 8":"BU","0,8H (8)":"BV","0,6H (8)":"BW","0,2H (8)":"BX","Abscisa 9":"BY","Prof 9":"BZ","0,8H (9)" :"CA","0,6H (9)":"CB","0,2H (9)":"CC","Abscisa 10":"CD","Prof 10":"CE","0,8H (10)":"CF","0,6H (10)":"CG","0,2H (10)":"CH","Abscisa 11":"CI","Prof 11":"CJ","0,8H (11)":"CK","0,6H (11)":"CL","0,2H (11)":"CM","Abscisa 12":"CN","Prof 12":"CO","0,8H (12)":"CP","0,6H (12)":"CQ","0,2H (12)":"CR","Abscisa 13":"CS","Prof 13":"CT","0,8H (13)":"CU","0,6H (13)":"CV","0,2H (13)":"CW","Abscisa 14":"CX","Prof 14":"CY","0,8H (14)":"CZ","0,6H (14)":"DA","0,2H (14)":"DB", "Coordenada_x":"T", "Coordenada_y":"T"}
celdas = ['D4',"J4","D5",'D6','D7','D8','G6','G7','M4','J9','A13','B13','C13','D13','E13','A14','B14','C14','D14','E14','A15','B15','C15','D15','E15','A16','B16','C16','D16','E16','A17','B17','C17','D17','E17','A18','B18','C18','D18','E18','A19','B19','C19','D19','E19','A20','B20','C20','D20','E20','A21','B21','C21','D21','E21','A22','B22','C22','D22','E22','A23','B23','C23','D23','E23','A24','B24','C24','D24','E24','A25','B25','C25','D25','E25','A26','B26','C26','D26','E26','G8','G9']

#for i in datos.keys():
#    print(i)

#print()
#print(len(datos))
#print()
#print(len(celdas))

datos_type = {"NomProy":'str',"CentCostos":'str',"Corriente":'str',"Fecha":'str',"Estación/cod":'str',"NomPunto":'str',"Dpto":'str',"Mpio":'str',"Responsables":'str',"AnchoT":'float',"Abscisa 1":'float',"Prof 1":'float',"0,8H (1)":'float',"0,6H (1)":'float',"0,2H (1)":'float',"Abscisa 2":'float',"Prof 2":'float',"0,8H (2)":'float',"0,6H (2)":'float',"0,2H (2)":'float',"Abscisa 3":'float',"Prof 3":'float',"0,8H (3)":'float',"0,6H (3)":'float',"0,2H (3)":'float',"Abscisa 4":'float',"Prof 4":'float',"0,8H (4)":'float',"0,6H (4)":'float',"0,2H (4)":'float',"Abscisa 5":'float',"Prof 5":'float',"0,8H (5)":'float',"0,6H (5)":'float',"0,2H (5)":'float',"Abscisa 6":'float',"Prof 6":'float',"0,8H (6)":'float',"0,6H (6)":'float',"0,2H (6)":'float',"Abscisa 7":'float',"Prof 7":'float',"0,8H (7)":'float',"0,6H (7)":'float',"0,2H (7)":'float',"Abscisa 8":'float',"Prof 8":'float',"0,8H (8)":'float',"0,6H (8)":'float',"0,2H (8)":'float',"Abscisa 9":'float',"Prof 9":'float',"0,8H (9)" :'float',"0,6H (9)":'float',"0,2H (9)":'float',"Abscisa 10":'float',"Prof 10":'float',"0,8H (10)":'float',"0,6H (10)":'float',"0,2H (10)":'float',"Abscisa 11":'float',"Prof 11":'float',"0,8H (11)":'float',"0,6H (11)":'float',"0,2H (11)":'float',"Abscisa 12":'float',"Prof 12":'float',"0,8H (12)":'float',"0,6H (12)":'float',"0,2H (12)":'float',"Abscisa 13":'float',"Prof 13":'float',"0,8H (13)":'float',"0,6H (13)":'float',"0,2H (13)":'float',"Abscisa 14":'float',"Prof 14":'float',"0,8H (14)":'float',"0,6H (14)":'float',"0,2H (14)":'float',"Coordenada_x":"str", "Coordenada_y":"str"}
nombre = 'file_creado'
for fila in range(2,10):
    
    # Crear una copia del archivo original
    libro_modificado = load_workbook(ruta_archivo_original)
    # Modificar el archivo copiado
    hoja_modificada = libro_modificado.active 

    #Dar los datos a las celdas
    lista_keys = list(datos.keys())             #Extraer una lista con las "claves" del diccionario
    for dato in range(len(datos)):
        key = lista_keys[dato]                  #Definir cada una de las claves, posición dato de la lista lista_keys
        pos_celda = datos[key] + str(fila)      #Obtener la posición de la celda y cocnatenar con el número de la fila
        valor = hoja_fuente[pos_celda].value
        
        if datos_type[key] == 'float':
            try:
                valor = float(valor)            #4.3 o 41,35 -- ['41','35']  -- 41.35
            except:
                try:
                    valor = valor.split(',')
                    valor = float(".".join(valor))
                except:
                    valor = ""
        
        if key == "Coordenada_x":
            valor = valor.split(" ")[0]
            
        if key == "Coordenada_y":
            valor = valor.split(" ")[1]    
                    
        hoja_modificada[celdas[dato]] = valor

    # Guardar el archivo Excel modificado en la nueva ruta
    libro_modificado.save(ruta_archivo_modificado+nombre+str(fila)+".xlsx")
    libro_modificado.close()
    
# Cerrar los archivos
libro_original.close()
# Esta prueba es para verificar el funcionamiento de GITHUB cuando se clona el repositorio
