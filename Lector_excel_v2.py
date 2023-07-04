from openpyxl import load_workbook
import openpyxl
# Especifica la ruta de tu archivo Excel
archivo_excel = 'D:/Desktop/Martín Anaya/03 Información Base/Solicitud_033.xlsx'

# Carga el archivo Excel
libro_excel = load_workbook(archivo_excel)

# Obtiene el nombre de la primera hoja
Nombres = libro_excel.sheetnames[0]

# Accede a la hoja por su nombre
hoja = libro_excel[Nombres]

# Itera sobre las filas de la hoja
#for fila in hoja.iter_rows():
#    for celda in fila:
#        # Imprime el valor de cada celda
#        print(celda.value)

hoja = libro_excel.active

#Inicia la variable para contar las filas
filas_llenas = 0

#Iteración sobre las filas y conteo
for fila in hoja.iter_rows():
    if not all(cell.value is None for cell in fila):
        filas_llenas += 1

print("Número de filas llenas:", filas_llenas)





