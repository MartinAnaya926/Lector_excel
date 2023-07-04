from openpyxl import load_workbook

# Ruta del archivo Excel original
ruta_archivo_original = 'D:/Desktop/Martín Anaya/02 VSC/01 Input/Prueba.xlsx'

# Ruta y nombre del archivo Excel modificado
ruta_archivo_modificado = "D:/Desktop/Martín Anaya/02 VSC/02 Output/Prueba_Salida_4.xlsx"

# Cargar el archivo Excel original
libro_original = load_workbook(ruta_archivo_original)

# Crear una copia del archivo original
libro_modificado = load_workbook(ruta_archivo_original)

# Modificar el archivo copiado
hoja_modificada = libro_modificado.active 

#Llenar celdas a partir de una lista dada "prueba"
prueba = ["Alacrán",12154656,"Río Cauca"]
celdas = ['A2',"B2","D2"]

print(len(prueba))
print()
print(len(celdas))

#Dar los datos a las celdas
for dato in range(len(prueba)):
    hoja_modificada[celdas[dato]] = prueba[dato]


# Guardar el archivo Excel modificado en la nueva ruta
libro_modificado.save(ruta_archivo_modificado)

# Cerrar los archivos
libro_original.close()
libro_modificado.close()

