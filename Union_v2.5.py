from openpyxl import load_workbook
import numpy as np

# Ruta del archivo que contiene los datos 
ruta_archivo_fuente = r"D:\Desktop\Martin_Anaya\01_Aforos_Liquidos\03_Informacion_Base\Solicitud_056.xlsx"
# Leer archivo fuente
libro_fuente = load_workbook(ruta_archivo_fuente)
# Modificar archivo fuente
hoja_fuente = libro_fuente.active
#Contador para el número de filas llenas
filas_llenas = 0

#Iteración sobre las y conteo de las filas llenas
for fila in hoja_fuente.iter_rows():
    if not all(cell is None for cell in fila):
        filas_llenas += 1
print()
print("Número de estaciones de aforo:", filas_llenas - 1)
print("Desde la número 2 hasta la número", filas_llenas)
print()

datos = {"Abscisa 1":"AK","Prof 1":"AL","0,6H (1)":"AN","Abscisa 2":"AP","Prof 2":"AQ","0,6H (2)":"AS","Abscisa 3":"AU","Prof 3":"AV","0,6H (3)":"AX","Abscisa 4":"AZ","Prof 4":"BA","0,6H (4)":"BC","Abscisa 5":"BE","Prof 5":"BF","0,6H (5)":"BH","Abscisa 6":"BJ","Prof 6":"BK","0,6H (6)":"BM","Abscisa 7":"BO","Prof 7":"BP","0,6H (7)":"BR","Abscisa 8":"BT","Prof 8":"BU","0,6H (8)":"BW","Abscisa 9":"BY","Prof 9":"BZ","0,6H (9)":"CB","Abscisa 10":"CD","Prof 10":"CE","0,6H (10)":"CG","Abscisa 11":"CI","Prof 11":"CJ","0,6H (11)":"CL","Abscisa 12":"CN","Prof 12":"CO","0,6H (12)":"CQ","Abscisa 13":"CS","Prof 13":"CT","0,6H (13)":"CV","Abscisa 14":"CX","Prof 14":"CY","0,6H (14)":"DA"}
datos1 = {"Código de estación":"K","Ancho total de la corriente":"L","Abscisa 1":"AK0","Prof 1":"AL","0,6H (1)":"AN","Abscisa 2":"AP","Prof 2":"AQ","0,6H (2)":"AS","Abscisa 3":"AU","Prof 3":"AV","0,6H (3)":"AX","Abscisa 4":"AZ","Prof 4":"BA","0,6H (4)":"BC","Abscisa 5":"BE","Prof 5":"BF","0,6H (5)":"BH","Abscisa 6":"BJ","Prof 6":"BK","0,6H (6)":"BM","Abscisa 7":"BO","Prof 7":"BP","0,6H (7)":"BR","Abscisa 8":"BT","Prof 8":"BU","0,6H (8)":"BW","Abscisa 9":"BY","Prof 9":"BZ","0,6H (9)":"CB","Abscisa 10":"CD","Prof 10":"CE","0,6H (10)":"CG","Abscisa 11":"CI","Prof 11":"CJ","0,6H (11)":"CL","Abscisa 12":"CN","Prof 12":"CO","0,6H (12)":"CQ","Abscisa 13":"CS","Prof 13":"CT","0,6H (13)":"CV","Abscisa 14":"CX","Prof 14":"CY","0,6H (14)":"DA"}

for fila in range(2,filas_llenas + 1):
    abscisas= []
    profundidades = []
    velocidades = []

    lista_keys = list(datos.keys())
    #print(lista_keys)
    c1 = 0
    c2 = 0
    #PUNTO SECO - PUNTO SECO - PUNTO SECO - PUNTO SECO - PUNTO SECO - PUNTO SECO - PUNTO SECO - PUNTO SECO - PUNTO SECO - PUNTO SECO 
    for dato in range(len(datos)):
        key = lista_keys[dato]
        celda = datos[key] + str(fila)
        valor = hoja_fuente[celda].value
        
        if valor is not None:
            c1 += 1
            
        else:
            c2 += 1
            
    if c1 == 0:
        print()
        print("¡¡¡¡¡¡¡¡¡¡EL PUNTO", str(fila), "ESTÁ SECO!!!!!!!!!!")
    else:
        print()        
    #    print("En el punto", str(fila), "faltan", c2, "datos")
    #    print("En el punto", str(fila), "hay", c1, "datos")

    #ALERTA FALTAN DATOS - ALERTA FALTAN DATOS - ALERTA FALTAN DATOS - ALERTA FALTAN DATOS - ALERTA FALTAN DATOS - ALERTA FALTAN DATOS - ALERTA FALTAN DATOS 
    lista_keys_1=list(datos1.keys())
    
    for dato in range(len(datos1)):
        
        key1 = lista_keys_1[dato]
        celda = datos1[key1] + str(fila)
        valor = hoja_fuente[celda].value

        a = 0
        b = 0
        c = 0 

        if key1 == "Código de estación":    
                nombre = valor
                print()
                print("NOMBRE DEL PUNTO NÚMERO",str(fila),":", nombre)

        if key1 == "Ancho total de la corriente":
            AnchoT = valor
            print("Ancho total del punto número",str(fila),":", AnchoT)
            print()
        
        
        for i in range (1,15):
             
            abs = "Abscisa " + str(i)
            prof = "Prof " + str(i)
            vel  = "0,6H (" + str(i) + ")"

            if key1 == abs:
                if valor is None:
                    print("No hay", abs)

                    celda_abs = datos1.get(key1) + str(fila)
                    abs_2_value = hoja_fuente[celda_abs].value
                    abscisas.append(abs_2_value)
                    
                else:
                    a = a + 1

                    celda_abs = datos1.get(key1) + str(fila)
                    abs_2_value = hoja_fuente[celda_abs].value
                    abscisas.append(abs_2_value)
                    #print(abscisas)

            if key1 == prof:
                if valor is None:
                    print("No hay", prof)

                    celda_prof = datos1.get(key1) + str(fila)
                    prof_2_value = hoja_fuente[celda_prof].value
                    profundidades.append(prof_2_value)

                else:
                    b = b + 1

                    celda_prof = datos1.get(key1) + str(fila)
                    prof_2_value = hoja_fuente[celda_prof].value
                    profundidades.append(prof_2_value)

            if key1 == vel:
                if valor is None:

                    celda_vel = datos1.get(key1) + str(fila)
                    vel_2_value = hoja_fuente[celda_vel].value
                    velocidades.append(vel_2_value)

                else:
                    b = b + 1

                    celda_vel = datos1.get(key1) + str(fila)
                    vel_2_value = hoja_fuente[celda_vel].value
                    velocidades.append(vel_2_value) 
                    
            if a==0 and b==0 and key1 == vel:
                if valor is None:
                    print("No hay velocidad en", vel)
                else:
                    c = c + 1
    print(abscisas)
    print(profundidades)
    print(velocidades)
    #VERIFICACIÓN DE ABSCISAS - VERIFICACIÓN DE ABSCISAS - VERIFICACIÓN DE ABSCISAS - VERIFICACIÓN DE ABSCISAS - VERIFICACIÓN DE ABSCISAS - VERIFICACIÓN DE ABSCISAS 

    #Los dos siguientes for separan los datos de la lista con comas y los unes con puntos para pasar de str a float
    for n,j in enumerate(abscisas):
        #print(n,j)
        try:
            float(abscisas[n])
        except:    
            try:
                abscisas[n] = j.split(',')
                abscisas[n] = float(".".join(abscisas[n]))
            except:
                abscisas[n] = abscisas[n]

    print()
    #print(abscisas)

    for n in reversed(range(len(abscisas))):
        if abscisas[n] == None:
            del abscisas[n]

    print("Abscisas:     ",abscisas)

    #Los dos siguientes for separan los datos de la lista profundidades con comas y los unes con puntos para pasar de str a float
    for p,q in enumerate(profundidades):
        #print(n,j)
        try:
            float(profundidades[p])
        except:    
            try:
                profundidades[p] = q.split(',')
                profundidades[p] = float(".".join(profundidades[p]))
            except:
                profundidades[p] = profundidades[p]

    #print(profundidades)

    for p in reversed(range(len(profundidades))):
        if profundidades[p] == None:
            del profundidades[p]
    print("Profundidades:",profundidades)

    #Los dos siguientes for separan los datos de la lista velocidades con comas y los unen con puntos para pasar de str a float
    for s,t in enumerate(velocidades):
        #print(s,t)
        try:
            float(velocidades[s])
        except:    
            try:
                velocidades[s] = t.split(',')
                velocidades[s] = float(".".join(velocidades[s]))
            except:
                velocidades[s] = velocidades[s]

    #print(velocidades)

    for s in reversed(range(len(velocidades))):
        if velocidades[s] == None:
            velocidades[s] = 0
    print("Velocidades:  ",velocidades)
    print()
    
    #COMPARACIONES
    #Los dos siguientes for comparan cada abscisa con la anterior y con el ancho total de la estación de aforo.
    z = 0
    for k in range(1,len(abscisas)):
        #print(abscisas[k])
        if k == len(abscisas) - 1:
            AnchoT_float = abscisas[k]

            #print("Número de abscisas con dato:", k +1)
            #print("Ancho total:", AnchoT_float)

        if abscisas[k] > abscisas[k-1]:
            z = z
        else:
            z = z + 1
            print("ALERTA: La Abscisa",k+1,"es menor que la abscisa",k)
            
    ancho = AnchoT_float
    if z > 0:
        print()
        print("Hay",z,"abscisas fuera de rango")

    for kk in range(len(abscisas)):
        if abscisas[kk] > ancho:
            print("La Abscisa",kk+1,"es mayor al ancho total")
    
    #Calcular el promedio y la desviación estandar de la lista profundidades
    if len(profundidades) > 0:
        prom_prof = round(sum(profundidades) / len(profundidades), 4)
        desest = round(np.std(profundidades), 4)
        #rango_i = round(prom_prof - 2*desest, 4)
        #rango_f = round(prom_prof + 2*desest, 4)
        #rango_i = round(-1*desest,4)
        #rango_f = round(desest,4)
        print()
        print("El rango aceptable para las profundidades es: [0, 1.3]")
        #print("El promedio de las profundidades en el punto",str(fila),"es:",prom_prof)
        #print("La Deviación Estándar de las profundidades en el punto",str(fila),"es:",desest)

    else: 
        print("No se puede calcular el promedio ni desviación estandar")

    for r in range(len(profundidades)):
        if 0 <= profundidades[r] < 1.3:
            continue
        else:
            print()
            print("ALERTA: La profundidad",r+1,"está fuera del rango aceptable")